from __future__ import annotations

import json
import re
from functools import lru_cache
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook
from common.assertions.sql_assertions import assert_sql_result


def _norm_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text in ("无", "None", "null"):
        return ""
    return text.replace("&#10;", "\n")


def _parse_json_obj(text: str) -> dict:
    text = _norm_text(text)
    if not text:
        return {}
    try:
        parsed = json.loads(text)
        if isinstance(parsed, dict):
            return parsed
    except Exception:
        # 兼容历史表格文本格式：
        # 1) Path参数userId={current_user_id}
        # 2) 可选参数: orderNo/state/goodName
        # 3) key=value,key2=value2
        lowered = text.replace("Path参数", "").replace("可选参数:", "").strip()
        pairs = re.findall(r"([A-Za-z_][A-Za-z0-9_]*)\s*=\s*([^,]+)", lowered)
        if pairs:
            return {k: v.strip() for k, v in pairs}
        keys = [k.strip() for k in re.split(r"[/,]", lowered) if re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", k.strip())]
        if keys:
            return {k: "" for k in keys}
        return {}
    return {}


def _parse_json_ex_data(text: str) -> dict[str, str]:
    text = _norm_text(text)
    if not text:
        return {}
    try:
        data = json.loads(text)
        if isinstance(data, dict):
            return {str(k): str(v) for k, v in data.items()}
    except Exception:
        pass
    return {}


def _render_text(template: str, context: dict[str, Any]) -> str:
    def repl(match: re.Match[str]) -> str:
        key = match.group(1).strip()
        return str(context.get(key, ""))

    return re.sub(r"\{\{\s*([^{}]+)\s*\}\}", repl, template)


def render_obj(value: Any, context: dict[str, Any]) -> Any:
    if isinstance(value, str):
        return _render_text(value, context)
    if isinstance(value, list):
        return [render_obj(v, context) for v in value]
    if isinstance(value, dict):
        return {k: render_obj(v, context) for k, v in value.items()}
    return value


@lru_cache(maxsize=1)
def _case_map() -> dict[str, dict]:
    files = sorted(Path(__file__).resolve().parents[2].glob("*.xlsx"))
    if not files:
        raise FileNotFoundError("No xlsx case file found in project root")

    wb = load_workbook(files[0], data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in ws[2]]
    rows: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if not row or not row[0]:
            continue
        item = {headers[i]: row[i] for i in range(len(headers))}
        case_id = str(item["id"]).strip()
        item["params_obj"] = _parse_json_obj(item.get("params"))
        item["data_obj"] = _parse_json_obj(item.get("data"))
        item["json_obj"] = _parse_json_obj(item.get("json"))
        item["json_ex_map"] = _parse_json_ex_data(item.get("jsonExData"))
        rows[case_id] = item
    return rows


def get_case(case_id: str) -> dict:
    data = _case_map()
    if case_id not in data:
        raise KeyError(f"Case not found in Excel: {case_id}")
    return data[case_id]


def build_request(case: dict, context: dict[str, Any]) -> tuple[str, dict, dict, dict]:
    path = _render_text(_norm_text(case.get("path")), context)
    params = render_obj(case.get("params_obj", {}), context) or {}
    data = render_obj(case.get("data_obj", {}), context) or {}
    json_body = render_obj(case.get("json_obj", {}), context) or {}

    # path 参数优先从 params 中替换（如 /api/address/{userId}）
    for k, v in list(params.items()):
        holder = "{" + str(k) + "}"
        if holder in path:
            path = path.replace(holder, str(v))
            params.pop(k, None)
    # 若 Excel params 为空，尝试用上下文直接替换 path 变量
    placeholders = re.findall(r"\{([A-Za-z_][A-Za-z0-9_]*)\}", path)
    for key in placeholders:
        if key in context:
            path = path.replace("{" + key + "}", str(context[key]))
        elif key == "userId" and "current_user_id" in context:
            path = path.replace("{userId}", str(context["current_user_id"]))
    return path, params, data, json_body


def assert_by_case_rule(response, case: dict) -> None:
    check = _norm_text(case.get("check"))
    expected = _norm_text(case.get("expected"))

    if not check:
        return

    if check == "res.json.code":
        options = [x.strip() for x in expected.split("|") if x.strip()]
        if not options:
            raise AssertionError(f"Invalid expected for {check}: {expected!r}")
        actual = None
        try:
            body = response.json()
            if isinstance(body, dict) and "code" in body:
                actual = str(body.get("code"))
        except Exception:
            actual = None
        # fallback: 非 JSON 响应时用 HTTP 状态码兜底
        if actual is None:
            actual = str(response.status_code)
        assert actual in options, f"Expected one of {options}, got {actual}, body={response.text}"
        return

    if check == "res.status_code":
        assert str(response.status_code) == expected, (
            f"Unexpected status code: {response.status_code}, expected={expected}, body={response.text}"
        )
        return

    if check == "res.text":
        assert expected in response.text, f"Expected {expected!r} in response text: {response.text}"
        return

    pytest.fail(f"Unsupported check rule in Excel: {check}")


def _pick_by_json_path(data: Any, path: str) -> Any:
    if path == "$":
        return data
    if not path.startswith("$."):
        return None

    cur: Any = data
    tokens = path[2:].split(".")
    for token in tokens:
        m = re.match(r"^([A-Za-z_][A-Za-z0-9_]*)(\[(\d+)\])?$", token)
        if not m:
            return None
        key = m.group(1)
        idx = m.group(3)
        if not isinstance(cur, dict) or key not in cur:
            return None
        cur = cur[key]
        if idx is not None:
            if not isinstance(cur, list):
                return None
            i = int(idx)
            if i >= len(cur):
                return None
            cur = cur[i]
    return cur


def extract_vars_from_response(case: dict, response_json: Any) -> dict[str, Any]:
    rules = case.get("json_ex_map") or {}
    if not isinstance(rules, dict):
        return {}
    result: dict[str, Any] = {}
    for key, path in rules.items():
        value = _pick_by_json_path(response_json, str(path))
        if value is not None:
            result[str(key)] = value
    return result


def run_sql_check(case: dict, context: dict[str, Any], db_client=None, db_tx=None) -> None:
    sql_tpl = _norm_text(case.get("sql_check"))
    expected = _norm_text(case.get("sql_expected"))
    if not sql_tpl or sql_tpl.upper() == "N/A" or not expected or expected.upper() == "N/A":
        return
    if db_client is None:
        return

    sql = _render_text(sql_tpl, context)
    actual = db_client.query_first(sql, conn=db_tx)
    assert_sql_result(actual, expected)
